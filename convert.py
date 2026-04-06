"""
Excelシフト管理ファイルをJSONに変換するスクリプト。
使い方: python convert.py シフト管理2026年04月度.xlsx [シフト管理2026年05月度.xlsx ...]
引数なしの場合、同じフォルダ内の「シフト管理*.xlsx」を全て変換。
出力: data/ フォルダに月別JSON + months.json (一覧)
"""
import openpyxl
import json
import glob
import sys
import os
import re
from pathlib import Path
from datetime import date
from collections import Counter

SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / "data"


def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    a1 = str(ws.cell(1, 1).value or "")
    m = re.search(r"(\d{4})\D+(\d{1,2})", a1)
    if not m:
        print(f"  スキップ: 年月を読み取れません ({a1})")
        return None
    year, month = int(m.group(1)), int(m.group(2))

    # Day numbers from row 1
    col_map = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, (int, float)) and 1 <= int(v) <= 31:
            col_map[int(v)] = c
    days = sorted(col_map.keys())

    # Day of week from row 2
    dow = []
    for d in days:
        v = ws.cell(2, col_map[d]).value
        dow.append(str(v) if v else "")

    # Events from row 3
    events = []
    for d in days:
        v = ws.cell(3, col_map[d]).value
        events.append(str(v).replace("\n", " ") if v else "")

    # Collect full names first to detect duplicates
    raw = []
    for r in range(4, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        name = str(v).strip()
        if name in ("行事予定", "打刻端末") or name.startswith("4/") or name.startswith("5/") or name.startswith("6/") or re.match(r"^\d+/", name):
            continue
        family = re.split(r"[\s\u3000]+", name)[0]
        raw.append((r, name, family))

    family_count = Counter(family for _, _, family in raw)
    family_seen = Counter()
    staff = []
    for r, full_name, family in raw:
        if family_count[family] > 1:
            parts = re.split(r"[\s\u3000]+", full_name)
            given = parts[1] if len(parts) > 1 else ""
            short = f"{family}({given[0]})" if given else family
        else:
            short = family

        shifts = {}
        for d in days:
            c = col_map[d]
            v = ws.cell(r, c).value
            if v is not None:
                shifts[str(d)] = str(v).replace("\n", " / ")

        if shifts:
            staff.append({"name": short, "shifts": shifts})

    # Start day of week (Mon=0)
    first = date(year, month, 1)
    js_dow = first.weekday()  # Mon=0 in Python

    days_in_month = (date(year, month % 12 + 1, 1) if month < 12 else date(year + 1, 1, 1)).toordinal() - first.toordinal()

    return {
        "year": year,
        "month": month,
        "daysInMonth": days_in_month,
        "days": days,
        "dow": dow,
        "events": events,
        "staff": staff,
        "startDow": js_dow,
    }


def main():
    DATA_DIR.mkdir(exist_ok=True)

    if len(sys.argv) > 1:
        files = sys.argv[1:]
    else:
        files = sorted(glob.glob(str(SCRIPT_DIR / "シフト管理*.xlsx"))) + \
                sorted(glob.glob(str(Path.home() / "Downloads" / "シフト管理*.xlsx")))

    if not files:
        print("対象ファイルが見つかりません。")
        print("使い方: python convert.py シフト管理2026年04月度.xlsx")
        return

    months = {}
    for f in files:
        print(f"変換中: {f}")
        data = parse_excel(f)
        if data:
            key = f"{data['year']}-{data['month']:02d}"
            out = DATA_DIR / f"{key}.json"
            with open(out, "w", encoding="utf-8") as fp:
                json.dump(data, fp, ensure_ascii=False)
            months[key] = f"{data['year']}年{data['month']}月"
            print(f"  -> {out}")

    # Write months index
    index_file = DATA_DIR / "months.json"
    existing = {}
    if index_file.exists():
        with open(index_file, "r", encoding="utf-8") as fp:
            existing = json.load(fp)
    existing.update(months)
    with open(index_file, "w", encoding="utf-8") as fp:
        json.dump(dict(sorted(existing.items())), fp, ensure_ascii=False, indent=2)
    print(f"\n完了！ {len(months)}件変換しました。")


if __name__ == "__main__":
    main()
